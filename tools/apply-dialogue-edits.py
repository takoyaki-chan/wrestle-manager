#!/usr/bin/env python3
"""
セリフ修正適用スクリプト
修正済みxlsxを読み込み、src/data.js の該当セリフを置換する。

使い方:
  python apply-dialogue-edits.py <edited.xlsx> [--dry-run]

  --dry-run: 実際の変更は行わず、差分一覧のみ表示

「全セリフ」シートでも性格別シートでも動作する。
修正案（最終列）が空でない行のみ処理する。
"""
import sys
import os

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl --break-system-packages")
    sys.exit(1)

def find_edits(xlsx_path):
    """xlsxから (現行セリフ, 修正案) のペアを抽出する"""
    wb = openpyxl.load_workbook(xlsx_path)
    edits = []

    for ws in wb.worksheets:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        # 「現行セリフ」と「修正案」の列を特定
        current_col = None
        revised_col = None
        for i, h in enumerate(headers):
            if h and '現行' in str(h):
                current_col = i + 1
            if h and '修正' in str(h):
                revised_col = i + 1

        if current_col is None or revised_col is None:
            continue

        for r in range(2, ws.max_row + 1):
            current = ws.cell(r, current_col).value
            revised = ws.cell(r, revised_col).value

            if not current or not revised:
                continue
            current = str(current).strip()
            revised = str(revised).strip()
            if current == revised:
                continue
            edits.append((current, revised))

    # 重複除去（複数シートに同じ修正がある場合）
    seen = set()
    unique = []
    for cur, rev in edits:
        key = (cur, rev)
        if key not in seen:
            seen.add(key)
            unique.append((cur, rev))

    return unique


def apply_edits(data_js_path, edits, dry_run=False):
    """data.js にセリフ置換を適用する"""
    with open(data_js_path, 'r', encoding='utf-8') as f:
        content = f.read()

    applied = []
    skipped = []
    conflicts = []

    for current, revised in edits:
        # JS文字列としての出現を検索（シングルクォート囲み）
        search_str = f"'{current}'"
        replace_str = f"'{revised}'"

        count = content.count(search_str)
        if count == 0:
            skipped.append((current, revised, "文字列が見つからない"))
        elif count > 1:
            conflicts.append((current, revised, f"{count}箇所に出現（手動確認が必要）"))
        else:
            if not dry_run:
                content = content.replace(search_str, replace_str, 1)
            applied.append((current, revised))

    if not dry_run and applied:
        with open(data_js_path, 'w', encoding='utf-8') as f:
            f.write(content)

    return applied, skipped, conflicts


def main():
    if len(sys.argv) < 2:
        print("使い方: python apply-dialogue-edits.py <edited.xlsx> [--dry-run]")
        print("  xlsx のあるディレクトリから wrestle-manager/src/data.js を探します")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    dry_run = '--dry-run' in sys.argv

    if not os.path.exists(xlsx_path):
        print(f"ファイルが見つかりません: {xlsx_path}")
        sys.exit(1)

    # data.js のパスを探す
    data_js_candidates = [
        'src/data.js',
        '../src/data.js',
        'wrestle-manager/src/data.js',
        os.path.join(os.path.dirname(xlsx_path), '..', 'src', 'data.js'),
    ]
    data_js_path = None
    for candidate in data_js_candidates:
        if os.path.exists(candidate):
            data_js_path = candidate
            break

    if not data_js_path:
        print("src/data.js が見つかりません。wrestle-manager リポジトリのルートで実行してください。")
        sys.exit(1)

    print(f"xlsx: {xlsx_path}")
    print(f"data.js: {data_js_path}")
    print(f"モード: {'ドライラン（変更なし）' if dry_run else '実適用'}")
    print()

    # 修正ペアを抽出
    edits = find_edits(xlsx_path)
    print(f"修正対象: {len(edits)}件")
    print()

    if not edits:
        print("修正案が入力されている行がありません。")
        sys.exit(0)

    # 適用
    applied, skipped, conflicts = apply_edits(data_js_path, edits, dry_run)

    # 結果表示
    if applied:
        print(f"{'[適用予定]' if dry_run else '[適用済み]'} {len(applied)}件:")
        for cur, rev in applied:
            print(f"  元: {cur}")
            print(f"  改: {rev}")
            print()

    if skipped:
        print(f"[スキップ] {len(skipped)}件:")
        for cur, rev, reason in skipped:
            print(f"  「{cur}」 → {reason}")
        print()

    if conflicts:
        print(f"[要手動確認] {len(conflicts)}件:")
        for cur, rev, reason in conflicts:
            print(f"  「{cur}」 → {reason}")
        print()

    # サマリー
    print("=" * 50)
    print(f"適用: {len(applied)}  スキップ: {len(skipped)}  要確認: {len(conflicts)}  合計: {len(edits)}")

    if dry_run and applied:
        print()
        print("実際に適用するには --dry-run を外して再実行してください。")


if __name__ == '__main__':
    main()
