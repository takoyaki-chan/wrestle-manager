#!/usr/bin/env python3
"""Generate full dialogue inventory Excel files from extracted JSON."""
import json, sys, os, subprocess
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Resolve paths ──
script_dir = os.path.dirname(os.path.abspath(__file__))
main_json_path = os.path.join(script_dir, '_all_lines.json')
battle_json_path = os.path.join(script_dir, '_battle_html_lines.json')

with open(main_json_path, 'r', encoding='utf-8') as f:
    data = json.load(f)
with open(battle_json_path, 'r', encoding='utf-8') as f:
    battle_html = json.load(f)

mainRows = data['mainRows']
battleRows = data['battleRows']
charMap = data['charMap']

OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'docs')

P_LABELS = {'normal':'ノーマル','bold':'強気','quiet':'寡黙','shy':'シャイ',
            'easygoing':'お気楽','earnest':'真面目','emotional':'感情的'}
A_LABELS = {'_default':'デフォルト','normal':'デフォルト',
            'polite':'丁寧','cool':'クール','ojousama':'お嬢様',
            'seductive':'蠱惑','delinquent':'不良','composed':'鷹揚'}

HEADER_FILL = PatternFill('solid', fgColor='4472C4')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

def style_header(ws, ncols):
    for col in range(1, ncols+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

# ══════════════════════════════════════════════════════════════
# 1. メインゲームセリフ一覧
# ══════════════════════════════════════════════════════════════
print(f"Main game rows: {len(mainRows)}")

wb = Workbook()

# ── Sheet 0: Summary ──
ws0 = wb.active
ws0.title = 'Summary'
ws0.append(['メインゲーム セリフ一覧'])
ws0.cell(row=1, column=1).font = Font(bold=True, size=14)
ws0.append([])

# Source-level summary
source_counts = {}
for r in mainRows:
    s = r['source']
    source_counts[s] = source_counts.get(s, 0) + 1
ws0.append(['出典', 'セリフ行数'])
style_header(ws0, 2)
# Move header styling to row 3
for col in range(1, 3):
    cell = ws0.cell(row=3, column=col)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
for src, cnt in sorted(source_counts.items(), key=lambda x: -x[1]):
    ws0.append([src, cnt])
ws0.append([])
ws0.append(['合計', len(mainRows)])
ws0.cell(row=ws0.max_row, column=1).font = Font(bold=True)
ws0.cell(row=ws0.max_row, column=2).font = Font(bold=True)
ws0.append([])
ws0.append(['生成日時', __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')])
ws0.column_dimensions['A'].width = 42
ws0.column_dimensions['B'].width = 14

# ── Sheet 1: 組合せ ──
ws1 = wb.create_sheet('組合せ')
ws1.append(['No.', '性格', '属性', '対象キャラ数', '対象キャラ'])
style_header(ws1, 5)

combo_no = 0
combos_seen = set()
for pKey in ['normal','bold','quiet','shy','easygoing','earnest','emotional']:
    for aKey in ['normal','polite','cool','ojousama','seductive','delinquent','composed']:
        key = f"{pKey}|{aKey}"
        chars = charMap.get(key, [])
        if not chars:
            continue
        combo_key = f"{pKey}_{aKey}"
        if combo_key in combos_seen:
            continue
        combos_seen.add(combo_key)
        combo_no += 1
        ws1.append([
            combo_no,
            P_LABELS.get(pKey, pKey),
            A_LABELS.get(aKey, aKey),
            len(chars),
            '、'.join(chars)
        ])

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 80

# ── Sheet 2: 全セリフ ──
ws2 = wb.create_sheet('全セリフ')
ws2.append(['No.', '出典', 'パス', '属性', '性格', '対象キャラ数', '対象キャラ', '現行セリフ', '書き直し案'])
style_header(ws2, 9)

for idx, row in enumerate(mainRows, 1):
    ws2.append([
        idx,
        row['source'],
        row['path'],
        row['archetype'],
        row['personality'],
        row['charCount'] if row['charCount'] > 0 else '',
        row['chars'],
        row['text'],
        '',
    ])

ws2.column_dimensions['A'].width = 7
ws2.column_dimensions['B'].width = 32
ws2.column_dimensions['C'].width = 28
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 14
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 50
ws2.column_dimensions['H'].width = 70
ws2.column_dimensions['I'].width = 70

# Auto-filter
ws2.auto_filter.ref = f"A1:I{len(mainRows)+1}"

out1 = os.path.join(OUT_DIR, 'dialogue-all-main.xlsx')
wb.save(out1)
print(f"Saved: {out1} ({len(mainRows)} lines)")

# ══════════════════════════════════════════════════════════════
# 2. 試合セリフ一覧
# ══════════════════════════════════════════════════════════════
wb2 = Workbook()

# ── Battle Summary ──
ws_sum2 = wb2.active
ws_sum2.title = 'Summary'
ws_sum2.append(['バトル系 セリフ一覧'])
ws_sum2.cell(row=1, column=1).font = Font(bold=True, size=14)
ws_sum2.append([])
battle_summary = [
    ('CUTIN_LINES (atk)', sum(len(v) if isinstance(v,list) else sum(len(x) for x in v.values() if isinstance(x,list)) for v in (battle_html.get('CUTIN_LINES',{}).get('atk',{})).values() if v)),
    ('CUTIN_LINES (def)', sum(len(v) if isinstance(v,list) else sum(len(x) for x in v.values() if isinstance(x,list)) for v in (battle_html.get('CUTIN_LINES',{}).get('def',{})).values() if v)),
    ('CUTIN_LINES (climax)', sum(len(v) if isinstance(v,list) else sum(len(x) for x in v.values() if isinstance(x,list)) for v in (battle_html.get('CUTIN_LINES',{}).get('climax',{})).values() if v)),
    ('CUTIN_LINES (bigmove)', sum(sum(len(x) for x in v.values() if isinstance(x,list)) for v in (battle_html.get('CUTIN_LINES',{}).get('bigmove',{})).values() if isinstance(v,dict))),
    ('DAMAGE_SERIF_LINES', sum(sum(len(x) for x in v.values() if isinstance(x,list)) for v in (battle_html.get('DAMAGE_SERIF_LINES',{})).values() if isinstance(v,dict))),
    ('DAMAGE_VOICE_LINES', sum(len(v) for v in (battle_html.get('DAMAGE_VOICE_LINES',{})).values() if isinstance(v,list))),
    ('VICTORY_LINES', len([r for r in battleRows if r['source']=='VICTORY_LINES'])),
    ('RIVALRY_MATCH_REACTION', len([r for r in battleRows if r['source']=='RIVALRY_MATCH_REACTION'])),
]
ws_sum2.append(['出典', 'セリフ行数'])
for col in range(1, 3):
    cell = ws_sum2.cell(row=3, column=col)
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = Alignment(horizontal='center')
for name, cnt in battle_summary:
    ws_sum2.append([name, cnt])
ws_sum2.append([])
ws_sum2.append(['合計', sum(c for _,c in battle_summary)])
ws_sum2.cell(row=ws_sum2.max_row, column=1).font = Font(bold=True)
ws_sum2.cell(row=ws_sum2.max_row, column=2).font = Font(bold=True)
ws_sum2.append([])
ws_sum2.append(['生成日時', __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')])
ws_sum2.column_dimensions['A'].width = 36
ws_sum2.column_dimensions['B'].width = 14

PERSONALITIES_ORDER = ['normal','bold','quiet','shy','easygoing','earnest','emotional']
ARCHETYPES_ORDER = ['normal','polite','seductive','delinquent','ojousama','cool','composed']
A_HEADERS = ['normal\n（普通）','polite\n（丁寧）','seductive\n（妖艶）','delinquent\n（ヤンキー）','ojousama\n（お嬢様）','cool\n（クール）','composed\n（鷹揚）']

def write_matrix_sheet(wb, title, obj, header_label):
    """Write a personality×archetype matrix sheet like battle-cutin-lines format."""
    ws = wb.create_sheet(title)

    ws.append([header_label] + [''] * len(ARCHETYPES_ORDER))
    ws.append(['性格 ＼ 口調'] + A_HEADERS)
    style_header(ws, len(ARCHETYPES_ORDER) + 1)
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)

    for pKey in PERSONALITIES_ORDER:
        pData = obj.get(pKey, {})
        if not pData:
            continue
        # Count chars for this personality
        pChars = []
        for k, v in charMap.items():
            if k.startswith(pKey + '|'):
                pChars.extend(v)
        pLabel = f"{pKey}\n（{P_LABELS.get(pKey, pKey)}）"

        row_data = [pLabel]
        for aKey in ARCHETYPES_ORDER:
            if isinstance(pData, dict):
                lines = pData.get(aKey, pData.get('_default', []))
            elif isinstance(pData, list):
                lines = pData
            else:
                lines = []
            if isinstance(lines, list):
                aChars = charMap.get(f"{pKey}|{aKey}", []) if aKey != 'normal' else charMap.get(f"{pKey}|normal", [])
                count = len(aChars)
                text = f"【{count}人】\n" + '\n'.join(str(l) for l in lines) if lines else ''
            else:
                text = ''
            row_data.append(text)
        ws.append(row_data)

    # Column widths
    ws.column_dimensions['A'].width = 16
    for i in range(len(ARCHETYPES_ORDER)):
        ws.column_dimensions[get_column_letter(i+2)].width = 35

    # Wrap text
    for row in ws.iter_rows(min_row=3, max_col=len(ARCHETYPES_ORDER)+1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

# ── CUTIN_LINES ──
cutin = battle_html.get('CUTIN_LINES', {})
for category in ['atk', 'def', 'climax', 'bigmove']:
    catData = cutin.get(category, {})
    if catData:
        cat_labels = {'atk':'カットイン（攻撃時）','def':'カットイン（防御時）','climax':'カットイン（クライマックス）','bigmove':'カットイン（ビッグムーブ）'}
        write_matrix_sheet(wb2, cat_labels.get(category, category), catData, cat_labels.get(category, category))

# ── DAMAGE_SERIF_LINES ──
dmg = battle_html.get('DAMAGE_SERIF_LINES', {})
if dmg:
    write_matrix_sheet(wb2, 'ダメージセリフ', dmg, 'ダメージセリフ（終盤・大技被弾時）')

# ── DAMAGE_VOICE_LINES ──
voice = battle_html.get('DAMAGE_VOICE_LINES', {})
if voice:
    ws_v = wb2.create_sheet('ダメージボイス')
    ws_v.append(['ダメージボイス（短い悲鳴・クライマックス被弾時）', ''])
    ws_v.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws_v.append(['口調', 'ボイス（改行区切り）'])
    style_header(ws_v, 2)
    for aKey in ['normal','polite','seductive','delinquent','ojousama','cool','composed']:
        lines = voice.get(aKey, [])
        if lines:
            label = f"{aKey}（{A_LABELS.get(aKey, aKey)}）"
            ws_v.append([label, '\n'.join(lines)])
    ws_v.column_dimensions['A'].width = 22
    ws_v.column_dimensions['B'].width = 60
    for row in ws_v.iter_rows(min_row=3, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

# ── VICTORY_LINES (per-character) ──
ws_vic = wb2.create_sheet('勝利セリフ（キャラ別）')
ws_vic.append(['キャラID', 'キャラ名', '性格', '属性', 'セリフNo.', '現行セリフ', '書き直し案'])
style_header(ws_vic, 7)
vic_rows = [r for r in battleRows if r['source'] == 'VICTORY_LINES']
for r in vic_rows:
    ws_vic.append([
        r['path'].split(' > ')[0],
        r['chars'],
        r['personality'],
        r['archetype'],
        r['path'].split(' > ')[-1],
        r['text'],
        '',
    ])
ws_vic.column_dimensions['A'].width = 10
ws_vic.column_dimensions['B'].width = 18
ws_vic.column_dimensions['C'].width = 12
ws_vic.column_dimensions['D'].width = 12
ws_vic.column_dimensions['E'].width = 10
ws_vic.column_dimensions['F'].width = 60
ws_vic.column_dimensions['G'].width = 60
ws_vic.auto_filter.ref = f"A1:G{len(vic_rows)+1}"

# ── RIVALRY_MATCH_REACTION (flat list) ──
rmr_rows = [r for r in battleRows if r['source'] == 'RIVALRY_MATCH_REACTION']
if rmr_rows:
    ws_rmr = wb2.create_sheet('試合中リアクション')
    ws_rmr.append(['No.', '出典', 'パス', '属性', '性格', '対象キャラ数', '対象キャラ', '現行セリフ', '書き直し案'])
    style_header(ws_rmr, 9)
    for idx, r in enumerate(rmr_rows, 1):
        ws_rmr.append([idx, r['source'], r['path'], r['archetype'], r['personality'],
                        r['charCount'] if r['charCount'] > 0 else '', r['chars'], r['text'], ''])
    ws_rmr.column_dimensions['H'].width = 60
    ws_rmr.auto_filter.ref = f"A1:I{len(rmr_rows)+1}"


out2 = os.path.join(OUT_DIR, 'dialogue-all-battle.xlsx')
wb2.save(out2)
print(f"Saved: {out2}")
print(f"  Victory lines: {len(vic_rows)}")
print(f"  RIVALRY_MATCH_REACTION: {len(rmr_rows)}")
print(f"  CUTIN categories: {list(cutin.keys())}")
print(f"  DAMAGE_SERIF personalities: {list(dmg.keys())}")
print(f"  DAMAGE_VOICE archetypes: {list(voice.keys())}")
print("Done!")
