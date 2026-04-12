import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('test/_session_f_data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = Workbook()
wb.remove(wb.active)

HEADERS = ['source', 'subcategory', 'event_id', 'personality', 'archetype',
           'outcome', 'dialogue_text', 'scene_text', 'notes']

HEADER_FILL = PatternFill('solid', fgColor='2C3E50')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
EMPTY_FILL = PatternFill('solid', fgColor='FFFDE7')  # light yellow for empty cells
EXISTING_FILL = PatternFill('solid', fgColor='E8F5E9')  # light green for existing
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='BDBDBD'),
    right=Side(style='thin', color='BDBDBD'),
    top=Side(style='thin', color='BDBDBD'),
    bottom=Side(style='thin', color='BDBDBD')
)

COL_WIDTHS = {'A': 28, 'B': 18, 'C': 12, 'D': 14, 'E': 14,
              'F': 14, 'G': 60, 'H': 60, 'I': 20}

PERSONALITY_JP = {
    'normal': 'ノーマル', 'bold': '強気', 'quiet': '寡黙', 'shy': '内気',
    'easygoing': 'お気楽', 'earnest': '真面目', 'emotional': '感情的'
}
ARCHETYPE_JP = {
    'ojousama': 'お嬢様', 'delinquent': 'ヤンキー', 'cool': 'クール',
    'seductive': '蠱惑', 'polite': '丁寧', 'composed': '鷹揚'
}

for source_name, rows in data.items():
    sheet_name = source_name[:31]  # Excel sheet name limit
    ws = wb.create_sheet(title=sheet_name)

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Column widths
    for col_letter, w in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = w

    # Headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # Data rows
    prev_subcat = None
    for row_idx, row_data in enumerate(rows, 2):
        values = [row_data.get(h, '') for h in HEADERS]
        is_empty = not row_data.get('dialogue_text') and not row_data.get('scene_text')
        is_new_subcat = row_data.get('subcategory') != prev_subcat
        prev_subcat = row_data.get('subcategory')

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=(col_idx >= 7))

            # Personality/archetype columns: show Japanese labels
            if col_idx == 4 and val in PERSONALITY_JP:
                cell.value = f'{PERSONALITY_JP[val]} ({val})'
            elif col_idx == 5 and val in ARCHETYPE_JP:
                cell.value = f'{ARCHETYPE_JP[val]} ({val})'

            # Highlight empty dialogue_text / scene_text
            if col_idx in (7, 8):
                if not val:
                    cell.fill = EMPTY_FILL
                else:
                    cell.fill = EXISTING_FILL

        # Bold subcategory on first row of group
        if is_new_subcat:
            ws.cell(row=row_idx, column=2).font = BOLD_FONT

    # Auto-filter
    ws.auto_filter.ref = f'A1:I{len(rows) + 1}'

# Summary sheet at the beginning
summary = wb.create_sheet('Summary', 0)
summary.freeze_panes = 'A2'
summary_headers = ['Source', 'Total Rows', 'Empty (to fill)', 'Existing', 'Sub-categories']
for col_idx, h in enumerate(summary_headers, 1):
    cell = summary.cell(row=1, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')
    cell.border = THIN_BORDER

for row_idx, (name, rows) in enumerate(data.items(), 2):
    empty = sum(1 for r in rows if not r.get('dialogue_text') and not r.get('scene_text'))
    existing = len(rows) - empty
    subcats = sorted(set(r.get('subcategory', '') for r in rows))
    vals = [name, len(rows), empty, existing, ', '.join(subcats)]
    for col_idx, v in enumerate(vals, 1):
        cell = summary.cell(row=row_idx, column=col_idx, value=v)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='top', wrap_text=(col_idx == 5))

summary.column_dimensions['A'].width = 32
summary.column_dimensions['B'].width = 12
summary.column_dimensions['C'].width = 14
summary.column_dimensions['D'].width = 12
summary.column_dimensions['E'].width = 80

# Total row
total_row = len(data) + 2
summary.cell(row=total_row, column=1, value='TOTAL').font = BOLD_FONT
for col in (2, 3, 4):
    cell = summary.cell(row=total_row, column=col)
    col_letter = get_column_letter(col)
    cell.value = f'=SUM({col_letter}2:{col_letter}{total_row - 1})'
    cell.font = BOLD_FONT
    cell.border = THIN_BORDER

wb.save('docs/dialogue-expansion-worksheet.xlsx')
print('Created docs/dialogue-expansion-worksheet.xlsx')
print(f'Total sheets: {len(wb.sheetnames)}')
for name in wb.sheetnames:
    print(f'  - {name}')
